import os
import django
import re
import requests
import json
from openpyxl import load_workbook

# Django init
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

from product import models

def clean_number(price):
    """Price ni integer ga aylantirish"""
    if price:
        num = re.sub(r"\D", "", str(price))
        return int(num) if num else 0
    return 0

def load_products_from_1c():
    """1C dan ma'lumot olib Product modelga saqlash"""
    file_path = "/app/WEB SAYT.xlsx"

    if not os.path.exists(file_path):
        print(f"❌ Fayl topilmadi: {file_path}")
        return

    # Excel fayldan artikullarni o'qish
    wb = load_workbook(file_path)
    sheet = wb.active

    article_list = [row[1] for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True) if row[1]]
    print(f"📌 Jami artikllar soni: {len(article_list)}")
    print("📌 Excel dan o'qilgan artikullar (10 ta namuna):", article_list[:10])

    if not article_list:
        print("❌ Excel faylda artikullar topilmadi.")
        return

    # 1C serveriga POST so'rov yuborish
    url = 'http://195.158.30.91/ONECOMPUTERS/hs/item/getdata'
    headers = {'Content-Type': 'application/json'}
    payload = {"sap_codes": article_list}

    try:
        response = requests.post(
            url,
            auth=('HttpUser', '85!@fdfd$DES35wgf&%'),
            headers=headers,
            data=json.dumps(payload),
            timeout=30
        )
    except Exception as e:
        print(f"❌ 1C serveriga so‘rov yuborishda xato: {e}")
        return

    if response.status_code != 200:
        print(f"❌ Serverdan noto‘g‘ri javob keldi: {response.status_code}")
        return

    try:
        data = response.json()
    except Exception as e:
        print(f"❌ JSON parse qilishda xato: {e}")
        return

    if "data" not in data:
        print(f"❌ JSON tarkibida 'data' yo‘q: {data}")
        return

    print("📌 1C serveridan kelgan data (10 ta namuna):", data["data"][:10])

    success_count = 0
    for product in data["data"]:
        try:
            item = product.get("Товар") or product.get("item")
            name = product.get("Наименование") or product.get("name")
            price = product.get("Цена") or product.get("price")
            remainder = product.get("Остатка") or product.get("quantity_left")

            if not item:
                print(f"⚠️ Item yo‘q, tashlab ketildi: {product}")
                continue

            clean_price = clean_number(price) if price else 0
            quantity_left = int(remainder) if remainder else 0

            # Product ni yangilash yoki yaratish
            obj, created = models.Product.objects.update_or_create(
                item=item,
                defaults={
                    "name": name or item,
                    "price": clean_price,
                    "quantity_left": max(0, quantity_left)
                }
            )

            print(f"{'CREATED' if created else 'UPDATED'}: {obj.item} | Name: {obj.name} | Price: {obj.price} | Qty: {obj.quantity_left}")
            success_count += 1

        except Exception as e:
            print(f"❌ Bitta productni saqlashda xato: {e} | Ma'lumot: {product}")

    print(f"✅ {success_count} ta mahsulot muvaffaqiyatli yangilandi.")


if __name__ == "__main__":
    load_products_from_1c()
